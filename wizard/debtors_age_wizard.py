# -*- coding: utf-8 -*-
import base64
import calendar
import io
import math
import os
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta

from odoo import fields, models, _
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None

from ..models.aging_engine import (
    calculate_debtor_ageing,
    normalize_header,
    parse_month,
)


BUILD_VERSION = "2026.09.02-r7-format"


class DclDebtorsAgeWizard(models.TransientModel):
    _name = "dcl.debtors.age.wizard"
    _description = "DCL Debtors Age Analysis"

    source_file = fields.Binary(string="Source Excel", required=True)
    source_filename = fields.Char(string="Source Filename")
    reporting_date = fields.Date(
        string="Reporting Date",
        required=True,
        default=fields.Date.context_today,
        help="The 'as at' date for the age analysis. The month/year should match the latest month in the source file.",
    )
    source_sheet = fields.Char(
        string="Source Sheet",
        help="Leave blank to auto-detect the worksheet containing the monthly debtors ledger.",
    )

    result_file = fields.Binary(string="Generated Excel", readonly=True)
    result_filename = fields.Char(string="Generated Filename", readonly=True)

    property_count = fields.Integer(string="Properties", readonly=True)
    debtor_count = fields.Integer(string="Debtors", readonly=True)
    exception_count = fields.Integer(string="Exceptions", readonly=True)
    total_receivables = fields.Monetary(string="Total Receivables", readonly=True)
    status_message = fields.Text(string="Status", readonly=True)
    currency_id = fields.Many2one(
        "res.currency",
        string="Currency",
        default=lambda self: self.env.company.currency_id,
        readonly=True,
    )

    # ------------------------------------------------------------------
    # Public actions
    # ------------------------------------------------------------------

    def action_generate(self):
        self.ensure_one()

        if not self.source_file:
            raise ValidationError(_("Please upload the source Debtors List Excel file."))
        if openpyxl is None:
            raise UserError(_("Python dependency 'openpyxl' is not installed."))
        if xlsxwriter is None:
            raise UserError(_("Python dependency 'xlsxwriter' is not installed."))

        workbook = None
        try:
            workbook = self._load_workbook()
            worksheet = self._select_sheet(workbook)
            parsed = self._read_source(worksheet)

            if not parsed["rows"]:
                raise UserError(
                    _("No debtor rows were found in worksheet '%s'.") % worksheet.title
                )

            self._validate_reporting_period(parsed["months"])

            # Release the uploaded workbook before rendering the output.
            try:
                workbook.close()
            finally:
                workbook = None

            result_bytes = self._build_xlsx(parsed)

            total_receivables = sum(
                float(row.get("current_outstanding") or 0.0)
                for row in parsed["rows"]
            )
            exception_count = sum(
                1 for row in parsed["rows"] if row.get("exception")
            )
            property_names = {
                row.get("property") or "Unassigned"
                for row in parsed["rows"]
            }

            report_date = self.reporting_date or fields.Date.today()
            filename = "Debtors_Age_Analysis_%s.xlsx" % report_date.strftime("%Y-%m-%d")

            self.write({
                "result_file": base64.b64encode(result_bytes),
                "result_filename": filename,
                "property_count": len(property_names),
                "debtor_count": len(parsed["rows"]),
                "exception_count": exception_count,
                "total_receivables": total_receivables,
                "status_message": _(
                    "Age analysis generated successfully from worksheet '%s'. "
                    "%s debtors across %s properties were processed. "
                    "%s reconciliation exceptions require review. Build %s."
                ) % (
                    parsed["sheet_name"],
                    len(parsed["rows"]),
                    len(property_names),
                    exception_count,
                    BUILD_VERSION,
                ),
            })

            return {
                "type": "ir.actions.act_window",
                "name": _("Debtors Age Analysis"),
                "res_model": self._name,
                "res_id": self.id,
                "view_mode": "form",
                "target": "new",
            }

        except (UserError, ValidationError):
            raise
        except Exception as exc:
            raise UserError(
                _("Unable to generate the age analysis.\n\n%s\n\nBuild: %s")
                % (str(exc), BUILD_VERSION)
            ) from exc
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

    def action_download(self):
        self.ensure_one()
        if not self.result_file:
            raise UserError(_("Generate the age analysis before downloading."))

        return {
            "type": "ir.actions.act_url",
            "url": (
                "/web/content?"
                "model=%s&id=%s&field=result_file&filename_field=result_filename"
                "&download=true"
            ) % (self._name, self.id),
            "target": "self",
        }

    # ------------------------------------------------------------------
    # Workbook loading / source discovery
    # ------------------------------------------------------------------

    def _load_workbook(self):
        payload = base64.b64decode(self.source_file)
        return openpyxl.load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=True,
        )

    @staticmethod
    def _parse_month_value(value):
        """Parse text, Excel dates, date objects and Excel serial month markers."""
        if value in (None, ""):
            return None

        if isinstance(value, datetime):
            return value.date().replace(day=1)
        if isinstance(value, date):
            return value.replace(day=1)

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            try:
                serial = float(value)
                if 20000 <= serial <= 80000:
                    parsed = date(1899, 12, 30) + timedelta(days=int(serial))
                    return parsed.replace(day=1)
            except (TypeError, ValueError, OverflowError):
                pass

        text = str(value).strip()
        if not text:
            return None

        cleaned = " ".join(
            text.replace(".", " ").replace(",", " ").split()
        )
        for fmt in (
            "%B %Y", "%b %Y", "%Y-%m", "%m/%Y",
            "%m-%Y", "%B-%Y", "%b-%Y",
        ):
            try:
                return datetime.strptime(cleaned, fmt).date().replace(day=1)
            except ValueError:
                continue

        try:
            parsed = parse_month(value)
            if parsed:
                if isinstance(parsed, datetime):
                    parsed = parsed.date()
                return parsed.replace(day=1)
        except Exception:
            pass

        return None

    @staticmethod
    def _row_values(ws, row_number):
        for row in ws.iter_rows(
            min_row=row_number,
            max_row=row_number,
            values_only=True,
        ):
            return tuple(row)
        return tuple()

    def _score_sheet(self, ws):
        first = self._row_values(ws, 1)
        second = self._row_values(ws, 2)
        score = 0.0

        for value in list(first) + list(second):
            if self._parse_month_value(value):
                score += 4

        expected = {
            "arrears_bf",
            "rent_levy",
            "recoveries",
            "adjustments",
            "receipts",
            "current_bal",
            "opening_balance",
            "total_billings",
            "total_receipts",
            "closing_balance",
        }
        for value in list(first) + list(second):
            if normalize_header(value) in expected:
                score += 2

        # Prefer the richer worksheet when two sheets contain the same ledger.
        score += min(ws.max_column or 0, 100) / 100.0
        score += min(ws.max_row or 0, 10000) / 10000.0
        return score

    def _select_sheet(self, workbook):
        requested = (self.source_sheet or "").strip()
        if requested:
            if requested not in workbook.sheetnames:
                raise UserError(
                    _("Worksheet '%s' does not exist. Available sheets: %s")
                    % (requested, ", ".join(workbook.sheetnames))
                )
            return workbook[requested]

        candidates = [(self._score_sheet(ws), ws) for ws in workbook.worksheets]
        candidates.sort(key=lambda item: item[0], reverse=True)
        if not candidates or candidates[0][0] <= 0:
            raise UserError(_("Could not identify the debtors ledger worksheet."))
        return candidates[0][1]

    # ------------------------------------------------------------------
    # Source parser
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_number(value):
        if value in (None, ""):
            return 0.0
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return 0.0
            return float(value)

        text = str(value).strip()
        if not text:
            return 0.0
        text = text.replace(",", "")
        text = text.replace("(", "-").replace(")", "")
        text = text.replace("KES", "").strip()
        try:
            return float(text)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _has_value(value):
        return value not in (None, "")

    @staticmethod
    def _text(value):
        return "" if value is None else str(value).strip()

    def _build_column_map(self, header1, header2):
        """
        Source layout confirmed against the supplied August workbook:
        row 1 = transaction headings; row 2 = month marker at each block start.
        """
        months = {}
        summary = {}
        current_month = None

        field_aliases = {
            "arrears_bf": "arrears",
            "rent_levy": "rent_levy",
            "recoveries": "recoveries",
            "adjustments": "adjustments",
            "receipts": "receipts",
            "current_bal": "current_bal",
            "current_balance": "current_bal",
        }
        summary_aliases = {
            "opening_balance": "opening_balance",
            "total_billings": "total_billings",
            "total_receipts": "total_receipts",
            "closing_balance": "closing_balance",
        }

        width = max(len(header1), len(header2))
        for idx in range(width):
            top = header1[idx] if idx < len(header1) else None
            bottom = header2[idx] if idx < len(header2) else None

            month = self._parse_month_value(bottom) or self._parse_month_value(top)
            if month:
                current_month = month.replace(day=1)
                months.setdefault(current_month, {})

            top_norm = normalize_header(top)
            bottom_norm = normalize_header(bottom)

            if top_norm in summary_aliases:
                summary[summary_aliases[top_norm]] = idx
            if bottom_norm in summary_aliases:
                summary[summary_aliases[bottom_norm]] = idx

            field_name = field_aliases.get(top_norm) or field_aliases.get(bottom_norm)
            if current_month and field_name:
                months.setdefault(current_month, {})[field_name] = idx

        required_fields = {
            "arrears", "rent_levy", "recoveries",
            "adjustments", "receipts", "current_bal",
        }
        valid_months = {}
        for month, columns in months.items():
            # A valid Dunhill month block should contain all six fields.
            if required_fields.issubset(columns.keys()):
                valid_months[month] = columns

        valid_months = dict(sorted(valid_months.items()))
        if not valid_months:
            debug_values = []
            for idx in range(width):
                top = header1[idx] if idx < len(header1) else None
                bottom = header2[idx] if idx < len(header2) else None
                if top not in (None, "") or bottom not in (None, ""):
                    debug_values.append(
                        "%s: [%r] / [%r]" % (idx + 1, top, bottom)
                    )
            raise UserError(
                _(
                    "Monthly columns could not be detected.\n\n"
                    "Detected headers:\n%s\n\nBuild: %s"
                ) % ("\n".join(debug_values[:80]), BUILD_VERSION)
            )

        return {"months": valid_months, "summary": summary}

    def _detect_identity_columns(self, header1, header2):
        combined = []
        width = max(len(header1), len(header2))
        for idx in range(width):
            combined.append(" ".join(
                part
                for part in [
                    normalize_header(header1[idx] if idx < len(header1) else None),
                    normalize_header(header2[idx] if idx < len(header2) else None),
                ]
                if part
            ))

        def find_index(words, default):
            for idx, text in enumerate(combined):
                if any(word in text for word in words):
                    return idx
            return default

        return {
            "account_id": find_index(
                ["debtor_id", "account_id", "customer_id", "tenant_id", "code"], 0
            ),
            "name": find_index(
                ["debtor_name", "customer_name", "tenant_name", "client_name", "tenant"], 1
            ),
            # Property heading is blank in the supplied workbook, so C is the safe fallback.
            "property": find_index(
                ["property", "site", "building", "estate", "project"], 2
            ),
        }

    def _read_source(self, ws):
        row_iterator = ws.iter_rows(values_only=True)
        try:
            header1 = tuple(next(row_iterator))
        except StopIteration:
            raise UserError(_("The selected worksheet is empty."))
        try:
            header2 = tuple(next(row_iterator))
        except StopIteration:
            header2 = tuple()

        column_map = self._build_column_map(header1, header2)
        month_blocks = column_map["months"]
        summary_columns = column_map["summary"]
        identity = self._detect_identity_columns(header1, header2)
        rows = []

        def cell(row_tuple, idx):
            if idx is None or idx < 0 or idx >= len(row_tuple):
                return None
            return row_tuple[idx]

        for row_tuple in row_iterator:
            account_id = self._text(cell(row_tuple, identity["account_id"]))
            debtor_name = self._text(cell(row_tuple, identity["name"]))
            property_name = self._text(cell(row_tuple, identity["property"]))

            if not account_id and not debtor_name:
                continue

            joined_identity = " ".join(
                [account_id.lower(), debtor_name.lower(), property_name.lower()]
            )
            if any(marker in joined_identity for marker in (
                "grand total", "portfolio total", "property total", "subtotal"
            )):
                continue

            monthly_rows = []
            for month, cols in month_blocks.items():
                monthly_rows.append({
                    "month": month,
                    "arrears_bf": self._safe_number(cell(row_tuple, cols["arrears"])),
                    "rent_levy": self._safe_number(cell(row_tuple, cols["rent_levy"])),
                    "recoveries": self._safe_number(cell(row_tuple, cols["recoveries"])),
                    "adjustments": self._safe_number(cell(row_tuple, cols["adjustments"])),
                    "receipts": self._safe_number(cell(row_tuple, cols["receipts"])),
                    "current_bal": self._safe_number(cell(row_tuple, cols["current_bal"])),
                })

            # The supplied enhanced sheet has summary headings BA:BD but the debtor
            # rows are blank there. Therefore, only use those cells when populated.
            raw_opening = cell(row_tuple, summary_columns.get("opening_balance"))
            raw_billings = cell(row_tuple, summary_columns.get("total_billings"))
            raw_receipts = cell(row_tuple, summary_columns.get("total_receipts"))
            raw_closing = cell(row_tuple, summary_columns.get("closing_balance"))

            source_closing = (
                self._safe_number(raw_closing)
                if self._has_value(raw_closing)
                else float(monthly_rows[-1]["current_bal"] or 0.0)
            )

            calculation = calculate_debtor_ageing(
                debtor_id=account_id,
                debtor_name=debtor_name,
                property_name=property_name or "Unassigned",
                monthly_rows=monthly_rows,
                source_closing=source_closing,
            )

            opening_balance = (
                self._safe_number(raw_opening)
                if self._has_value(raw_opening)
                else calculation.opening_balance
            )
            period_charges = (
                self._safe_number(raw_billings)
                if self._has_value(raw_billings)
                else calculation.period_charges
            )
            # Source receipts are stored as negative movements. The report displays
            # collections as a positive amount, matching the reference report label.
            period_receipts = (
                self._safe_number(raw_receipts)
                if self._has_value(raw_receipts)
                else -calculation.period_receipts
            )

            reconciliation_difference = (
                float(calculation.calculated_outstanding or 0.0)
                - float(calculation.current_outstanding or 0.0)
            )

            rows.append({
                "account_id": calculation.debtor_id,
                "debtor_name": calculation.debtor_name,
                "property": calculation.property_name or "Unassigned",
                "billing_frequency": calculation.billing_frequency,
                "opening_balance": opening_balance,
                "period_charges": period_charges,
                "period_receipts": period_receipts,
                "current_outstanding": calculation.current_outstanding,
                "ageing": calculation.aged,
                "reconciliation_difference": reconciliation_difference,
                "exception": bool(calculation.exception),
                "exception_message": calculation.exception or "",
                # Retained so the final property report can reproduce the
                # reference workbook's month-by-month balance snapshots.
                "monthly_rows": monthly_rows,
            })

        return {
            "rows": rows,
            "months": list(month_blocks.keys()),
            "sheet_name": ws.title,
        }

    def _validate_reporting_period(self, months):
        if not months or not self.reporting_date:
            return
        latest = max(months)
        if (latest.year, latest.month) != (
            self.reporting_date.year,
            self.reporting_date.month,
        ):
            raise ValidationError(
                _(
                    "Reporting Date does not match the latest source month. "
                    "The source ends in %s, but the Reporting Date is %s."
                ) % (
                    latest.strftime("%B %Y"),
                    self.reporting_date.strftime("%d %B %Y"),
                )
            )

    # ------------------------------------------------------------------
    # XLSX generation
    # ------------------------------------------------------------------

    @staticmethod
    def _safe_sheet_name(name, used):
        invalid = '[]:*?/\\'
        cleaned = "".join(
            "_" if char in invalid else char for char in (name or "Property")
        ).strip() or "Property"
        cleaned = cleaned[:31]
        base = cleaned
        suffix = 1
        while cleaned.lower() in used:
            suffix += 1
            tag = "_%s" % suffix
            cleaned = (base[: 31 - len(tag)] + tag)[:31]
        used.add(cleaned.lower())
        return cleaned

    @staticmethod
    def _ordinal(day):
        if 10 <= day % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return "%s%s" % (day, suffix)

    @staticmethod
    def _month_end(month_date):
        return date(
            month_date.year,
            month_date.month,
            calendar.monthrange(month_date.year, month_date.month)[1],
        )

    def _debtor_snapshots(self, debtor, source_months):
        """
        Reproduce the reference workbook's per-debtor snapshot rows.

        For each source month, calculate ageing using only transactions up to
        that month. The current month uses the reporting date; prior months use
        month-end dates. Balance is the source Current Bal for that period.
        """
        monthly_rows = sorted(
            list(debtor.get("monthly_rows") or []),
            key=lambda item: item["month"],
        )

        if not monthly_rows:
            return []

        snapshots = []

        for idx in range(len(monthly_rows) - 1, -1, -1):
            cutoff_row = monthly_rows[idx]
            subset = monthly_rows[: idx + 1]

            result = calculate_debtor_ageing(
                debtor_id=debtor.get("account_id", ""),
                debtor_name=debtor.get("debtor_name", ""),
                property_name=debtor.get("property", ""),
                monthly_rows=subset,
                source_closing=cutoff_row.get("current_bal"),
            )

            cutoff_month = cutoff_row["month"]
            snapshot_date = (
                self.reporting_date
                if idx == len(monthly_rows) - 1
                else self._month_end(cutoff_month)
            )

            snapshots.append({
                "date": snapshot_date,
                "balance": float(cutoff_row.get("current_bal") or 0.0),
                "aged": result.aged,
                "exception": bool(result.exception),
                "exception_message": result.exception or "",
            })

        return snapshots

    def _build_xlsx(self, parsed):
        rows = parsed["rows"]
        source_months = sorted(parsed["months"], reverse=True)
        chronological_months = sorted(parsed["months"])

        grouped = defaultdict(list)
        for row in rows:
            grouped[row.get("property") or "Unassigned"].append(row)

        fd, path = tempfile.mkstemp(prefix="dcl_debtors_age_", suffix=".xlsx")
        os.close(fd)

        try:
            workbook = xlsxwriter.Workbook(path)

            navy = "#1E2761"
            navy_2 = "#2A357A"
            pale = "#F2F4F8"
            pale_2 = "#E8EDF5"
            white = "#FFFFFF"
            green = "#92D050"

            money_2dp = '#,##0.00;[Red](#,##0.00);-'
            money_0dp = '#,##0;[Red]-#,##0;"-"'
            date_fmt_code = "yyyy-mm-dd"

            fmt_master_title = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 16,
                "bold": True,
                "font_color": white,
                "bg_color": navy,
                "align": "center",
                "valign": "vcenter",
            })
            fmt_master_subtitle = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "font_color": navy,
                "bg_color": pale,
                "align": "center",
                "valign": "vcenter",
            })
            fmt_kpi = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 11,
                "bold": True,
                "font_color": navy,
                "bg_color": pale,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "border": 1,
                "border_color": pale_2,
            })
            fmt_summary_header = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": white,
                "bg_color": navy_2,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "border": 1,
                "border_color": white,
            })
            fmt_summary_text_odd = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": white,
                "border": 1,
                "border_color": pale_2,
            })
            fmt_summary_text_even = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": pale,
                "border": 1,
                "border_color": pale_2,
            })
            fmt_summary_int_odd = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": white,
                "align": "right",
                "border": 1,
                "border_color": pale_2,
                "num_format": "#,##0",
            })
            fmt_summary_int_even = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": pale,
                "align": "right",
                "border": 1,
                "border_color": pale_2,
                "num_format": "#,##0",
            })
            fmt_summary_money_odd = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": white,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })
            fmt_summary_money_even = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": pale,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })
            fmt_summary_current_odd = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": navy,
                "bg_color": white,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })
            fmt_summary_current_even = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": navy,
                "bg_color": pale,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })
            fmt_summary_green_text = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": green,
                "border": 1,
                "border_color": pale_2,
            })
            fmt_summary_green_int = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": green,
                "border": 1,
                "border_color": pale_2,
                "num_format": "#,##0",
            })
            fmt_summary_green_money = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bg_color": green,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })
            fmt_summary_green_current = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": navy,
                "bg_color": green,
                "border": 1,
                "border_color": pale_2,
                "num_format": money_2dp,
            })

            fmt_property_total = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 11,
                "bold": True,
                "font_color": navy,
                "bg_color": pale,
                "align": "center",
                "valign": "vcenter",
            })
            fmt_property_header = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
            })
            fmt_property_month_header = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "bg_color": pale,
                "align": "center",
            })
            fmt_property_name = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": navy,
            })
            fmt_property_date = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 11,
                "num_format": date_fmt_code,
            })
            fmt_property_balance = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 10,
                "bold": True,
                "font_color": navy,
                "num_format": money_0dp,
            })
            fmt_property_money = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 11,
                "num_format": money_0dp,
            })
            fmt_property_frequency = workbook.add_format({
                "font_name": "Calibri",
                "font_size": 11,
            })

            # ---------------- Portfolio Summary ----------------
            summary = workbook.add_worksheet("Portfolio Summary")
            summary.hide_gridlines(2)

            total_receivables = sum(
                float(item.get("current_outstanding") or 0.0)
                for item in rows
            )
            total_collections = sum(
                float(item.get("period_receipts") or 0.0)
                for item in rows
            )
            total_charges = sum(
                float(item.get("period_charges") or 0.0)
                for item in rows
            )
            collection_efficiency = (
                total_collections / total_charges
                if abs(total_charges) > 0.005
                else 0.0
            )

            summary.merge_range(
                "A1:S1",
                "PORTFOLIO DEBTORS AGING & RECONCILIATION MASTER SUMMARY",
                fmt_master_title,
            )
            summary.merge_range(
                "A2:S2",
                "Reporting Date: %s  |  Currency: %s  |  Total Properties Monitored: %s"
                % (
                    self.reporting_date.strftime("%d %B %Y"),
                    self.currency_id.name or "KES",
                    len(grouped),
                ),
                fmt_master_subtitle,
            )
            summary.set_row(0, 28)
            summary.set_row(1, 22)

            summary.merge_range(
                "B4:D5",
                "Total Receivables\n%s %s"
                % (self.currency_id.name or "KES", f"{total_receivables:,.2f}"),
                fmt_kpi,
            )
            summary.merge_range(
                "F4:H5",
                "Monitored Sites\n%s Properties" % len(grouped),
                fmt_kpi,
            )
            summary.merge_range(
                "J4:L5",
                "Total Active Debtors\n%s Debtors" % f"{len(rows):,}",
                fmt_kpi,
            )
            summary.merge_range(
                "N4:P5",
                "Total Period Collections\n%s %s"
                % (self.currency_id.name or "KES", f"{total_collections:,.2f}"),
                fmt_kpi,
            )
            summary.merge_range(
                "R4:S5",
                "Collection Efficiency\n%s" % f"{collection_efficiency:.1%}",
                fmt_kpi,
            )

            earliest = min(chronological_months)
            previous_month_end = earliest - timedelta(days=1)
            period_label = "%s-%s" % (
                chronological_months[0].strftime("%b"),
                chronological_months[-1].strftime("%b"),
            )

            summary_headers = [
                "#",
                "Property / Site Name",
                "Debtors",
                "Opening Arrears (%s)" % previous_month_end.strftime("%b %y"),
                "Period Charges (%s)" % period_label,
                "Period Receipts (%s)" % period_label,
                "Current Outstanding",
            ]
            summary_headers += [m.strftime("%b %Y") for m in source_months]
            summary_headers += ["Opening", "Exceptions"]

            header_row = 6
            for col, heading in enumerate(summary_headers):
                summary.write(header_row, col, heading, fmt_summary_header)
            summary.set_row(header_row, 36)
            summary.freeze_panes(header_row + 1, 2)

            out_row = header_row + 1

            for seq, property_name in enumerate(
                sorted(grouped, key=lambda x: x.lower()),
                start=1,
            ):
                property_rows = grouped[property_name]
                aged_totals = defaultdict(float)

                for debtor in property_rows:
                    for key, value in (debtor.get("ageing") or {}).items():
                        aged_totals[key] += float(value or 0.0)

                item = {
                    "property": property_name,
                    "debtors": len(property_rows),
                    "opening": sum(float(x.get("opening_balance") or 0.0) for x in property_rows),
                    "charges": sum(float(x.get("period_charges") or 0.0) for x in property_rows),
                    "receipts": sum(float(x.get("period_receipts") or 0.0) for x in property_rows),
                    "current": sum(float(x.get("current_outstanding") or 0.0) for x in property_rows),
                    "aged": aged_totals,
                    "exceptions": sum(1 for x in property_rows if x.get("exception")),
                }

                has_exception = item["exceptions"] > 0
                even = seq % 2 == 0

                if has_exception:
                    text_fmt = fmt_summary_green_text
                    int_fmt = fmt_summary_green_int
                    money_fmt = fmt_summary_green_money
                    current_fmt = fmt_summary_green_current
                else:
                    text_fmt = fmt_summary_text_even if even else fmt_summary_text_odd
                    int_fmt = fmt_summary_int_even if even else fmt_summary_int_odd
                    money_fmt = fmt_summary_money_even if even else fmt_summary_money_odd
                    current_fmt = fmt_summary_current_even if even else fmt_summary_current_odd

                summary.write_number(out_row, 0, seq, int_fmt)
                summary.write(out_row, 1, item["property"], text_fmt)
                summary.write_number(out_row, 2, item["debtors"], int_fmt)
                summary.write_number(out_row, 3, item["opening"], money_fmt)
                summary.write_number(out_row, 4, item["charges"], money_fmt)
                summary.write_number(out_row, 5, item["receipts"], money_fmt)
                summary.write_number(out_row, 6, item["current"], current_fmt)

                col = 7
                for month in source_months:
                    summary.write_number(
                        out_row,
                        col,
                        float(item["aged"].get(month.strftime("%b %Y"), 0.0) or 0.0),
                        money_fmt,
                    )
                    col += 1

                summary.write_number(
                    out_row,
                    col,
                    float(item["aged"].get("Opening", 0.0) or 0.0),
                    money_fmt,
                )
                col += 1
                summary.write_number(out_row, col, item["exceptions"], int_fmt)
                out_row += 1

            summary.autofilter(header_row, 0, out_row - 1, len(summary_headers) - 1)
            summary.set_column("A:A", 5)
            summary.set_column("B:B", 28)
            summary.set_column("C:C", 10)
            summary.set_column("D:G", 20)
            summary.set_column("H:O", 14)
            summary.set_column("P:P", 15)
            summary.set_column("Q:Q", 12)

            # ---------------- Property sheets ----------------
            used_sheet_names = {"portfolio summary"}
            final_month_labels = [
                month.strftime("%b %Y") for month in source_months
            ]

            current_header = "%s %s" % (
                self.reporting_date.strftime("%B"),
                self._ordinal(self.reporting_date.day),
            )
            visible_age_headers = [current_header]
            visible_age_headers += [
                month.strftime("%B") for month in source_months[1:]
            ]
            visible_age_headers += [
                "%s %s" % (
                    chronological_months[0].strftime("%B"),
                    self._ordinal(1),
                )
            ]

            for property_name in sorted(grouped, key=lambda x: x.lower()):
                property_rows = grouped[property_name]
                sheet_name = self._safe_sheet_name(property_name, used_sheet_names)
                ws = workbook.add_worksheet(sheet_name)
                ws.hide_gridlines(2)

                ws.merge_range("B1:K1", "Total Outstanding", fmt_property_total)

                ws.write("A2", "Tenant / Date", fmt_property_header)
                ws.write("B2", "Balance", fmt_property_header)

                for idx, heading in enumerate(visible_age_headers):
                    ws.write(1, 2 + idx, heading, fmt_property_month_header)

                ws.write("M2", "Billing Frequency", fmt_property_header)
                ws.freeze_panes(2, 2)

                ws.set_column("A:A", 22)
                ws.set_column("B:B", 17)
                ws.set_column("C:D", 14)
                ws.set_column("E:J", 12)
                ws.set_column("K:K", 13)
                ws.set_column("L:L", 3)
                ws.set_column("M:M", 18)

                out_row = 2

                for debtor in property_rows:
                    debtor_name = debtor.get("debtor_name", "")
                    frequency = debtor.get("billing_frequency", "")
                    account_id = debtor.get("account_id", "")

                    ws.write(out_row, 0, debtor_name, fmt_property_name)
                    if account_id:
                        ws.write_comment(
                            out_row,
                            0,
                            "Account ID: %s" % account_id,
                            {"author": "DCL Debtors Age Analysis"},
                        )
                    ws.write(out_row, 12, frequency, fmt_property_frequency)

                    for col in range(2, 11):
                        ws.write_blank(out_row, col, None, fmt_property_money)

                    snapshots = self._debtor_snapshots(debtor, source_months)

                    for snapshot in snapshots:
                        out_row += 1
                        ws.write_datetime(
                            out_row,
                            0,
                            datetime.combine(snapshot["date"], datetime.min.time()),
                            fmt_property_date,
                        )
                        ws.write_number(
                            out_row,
                            1,
                            snapshot["balance"],
                            fmt_property_balance,
                        )

                        aged = snapshot["aged"] or {}

                        col = 2
                        for month_label in final_month_labels:
                            ws.write_number(
                                out_row,
                                col,
                                float(aged.get(month_label, 0.0) or 0.0),
                                fmt_property_money,
                            )
                            col += 1

                        ws.write_number(
                            out_row,
                            10,
                            float(aged.get("Opening", 0.0) or 0.0),
                            fmt_property_money,
                        )
                        ws.write(out_row, 12, frequency, fmt_property_frequency)

                        if snapshot.get("exception_message"):
                            ws.write_comment(
                                out_row,
                                1,
                                snapshot["exception_message"],
                                {"author": "DCL Debtors Age Analysis"},
                            )

                    out_row += 5

                ws.set_landscape()
                ws.fit_to_pages(1, 0)
                ws.repeat_rows(0, 1)
                ws.set_margins(left=0.25, right=0.25, top=0.4, bottom=0.4)

            workbook.close()

            with open(path, "rb") as handle:
                return handle.read()

        finally:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass
